"""Generate the WTG-catalog-upload template xlsx.

Produces `WTG_template.xlsx` matching the format BESSTY's in-app
"Upload xlsx" button accepts (the WTG-style workbook parsed by
`web/src/lib/xlsxImport.ts::parseWtgWorkbook`):

  - One workbook per turbine model.
  - One sheet per operating mode.
  - Per-sheet layout:
      A1 "Model:"   B1 <model name>
      A2 "Mode:"    B2 <mode name>
      A3 "Type:"    B3 "WTG"
      A4 (blank)    B4..  wind speeds in m/s @ 10 m (one column each)
      A5..          frequencies in Hz (rows), B5.. Lw values per wind speed

Includes:
  - A README sheet explaining the layout.
  - 3 example modes with octave-band sound power.
  - Plausible but obviously-placeholder Lw values the user replaces with
    their own datasheet numbers.

Re-run any time the format changes. The output file lives alongside this
script so it can be downloaded directly by a user from the repo's
Examples/ folder.
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from pathlib import Path

OUT = Path(__file__).parent / "WTG_template.xlsx"

# Octave bands (16 Hz–8 kHz) — match the BESSTY solver's 10-band octave system.
FREQS_HZ = [16, 31.5, 63, 125, 250, 500, 1000, 2000, 4000, 8000]

# Wind speeds @ 10 m AGL — typical IEC 61400-11 reporting range.
WIND_SPEEDS = [3, 4, 5, 6, 7, 8, 9, 10, 11, 12]

# Example modes for the template. Replace the Lw arrays with the
# manufacturer's per-band, per-wind-speed data.
#
# `lw_at_8ms`: a flat starting per-band spectrum at 8 m/s. The script
# fills the rest of the grid by scaling slightly with wind speed so the
# template looks realistic rather than uniform. Replace with real data.
MODES = [
    {
        "name": "Nominal",
        "lw_at_8ms": [80.0, 88.0, 96.0, 100.0, 103.0, 105.0, 103.0, 100.0, 95.0, 89.0],
    },
    {
        "name": "NoiseReduced-3dB",
        "lw_at_8ms": [77.0, 85.0, 93.0,  97.0, 100.0, 102.0, 100.0,  97.0, 92.0, 86.0],
    },
    {
        "name": "NoiseReduced-6dB",
        "lw_at_8ms": [74.0, 82.0, 90.0,  94.0,  97.0,  99.0,  97.0,  94.0, 89.0, 83.0],
    },
]


def lw_for(spec_at_8ms, wind_speed, band_idx):
    """Plausible scaling — Lw rises ~+1 dB per m/s above 4 m/s, flatlines
    above ~10 m/s. Bands above 2 kHz lose a bit more at low wind speeds
    because they're tip-noise dominated."""
    base = spec_at_8ms[band_idx]
    if wind_speed < 4:
        return None      # cut-in below 4 m/s
    if wind_speed <= 10:
        bump = (wind_speed - 8) * 1.0
    else:
        bump = 2.0       # plateau
    # High-frequency rolloff at low winds.
    if wind_speed < 7 and band_idx >= 7:
        bump -= 1.5
    return round(base + bump, 1)


def write_readme(ws):
    ws.title = "README"
    ws.column_dimensions["A"].width = 110
    bold = Font(bold=True, size=13)
    body = Font(size=11)
    ws["A1"] = "BESSTY — WTG catalog template (read me first)"
    ws["A1"].font = bold
    lines = [
        "",
        "How to use this file",
        "1. Duplicate this file once per turbine MODEL you want to add.",
        "2. Rename it — the file name is just a label, but a clear name (e.g. 'V163-4.5MW.xlsx') helps.",
        "3. Edit the MODEL name in cell B1 of each mode sheet — must be the same on every sheet.",
        "4. Edit the MODE name in cell B2 of each sheet — one mode per sheet (PO4500, NRO+0, etc.).",
        "5. Edit the wind-speed header row (row 4 columns B onwards) — m/s @ 10 m AGL.",
        "6. Edit the frequency column (col A from row 5 onwards) — Hz, ascending.",
        "7. Fill the grid (B5 onwards) with the per-band, per-wind-speed sound-power level in dB.",
        "8. Leave a cell blank for any wind speed below cut-in (BESSTY treats blank as 0).",
        "",
        "Then in BESSTY:",
        "  Catalog → ↑ Upload xlsx → pick this file → choose A-weighted or Z-weighted.",
        "  - Z-weighted (un-weighted) = ISO 9613-2 convention. Use this when the datasheet shows the un-weighted Lw per band, with the overall LwA totalled separately.",
        "  - A-weighted (LwA per band) = IEC 61400-11 convention. BESSTY un-weights internally before propagation.",
        "  Picking wrong → propagated levels off by ~3-5 dB.",
        "",
        "Layout reference (per sheet)",
        "  Row 1: 'Model:' | <model name>          ← BESSTY uses this as the displayed name.",
        "  Row 2: 'Mode:'  | <mode name>           ← shown in the Source's mode dropdown.",
        "  Row 3: 'Type:'  | 'WTG'                 ← keep as WTG.",
        "  Row 4: (blank)  | 3 | 4 | 5 | ...       ← wind speeds, one column each, m/s @ 10 m.",
        "  Row 5+: frequency (Hz) | Lw at 3 m/s | Lw at 4 m/s | ...   ← one row per band.",
        "",
        "Frequency bands",
        "  This template uses octave bands (16 Hz – 8 kHz, 10 bands).",
        "  To use one-third octave instead, replace the freq column with the 31-band 1/3-octave set",
        "  (10, 12.5, 16, 20, 25, 31.5, ..., 8000, 10000) — BESSTY auto-detects the band system from the spacing.",
        "",
        "Per-sheet sanity check",
        "  Every sheet in this workbook must have the same model name (B1) and same frequency set (col A).",
        "  Modes can use different wind-speed ranges if needed (BESSTY interpolates by exact match — so",
        "  the wind speed you set in the project's Scenario tab needs to appear as a column header).",
        "",
        "If your data is per-A-weighted-band only (no per-Z-band):",
        "  Tick the A-weighted option when uploading. BESSTY converts to Z via the IEC 61672-1 offsets.",
        "",
        "Delete this README sheet before uploading? No need — BESSTY ignores it (it only reads sheets",
        "  whose row 1 column A is 'Model:').",
    ]
    for i, t in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=t)
        c.font = body
        c.alignment = Alignment(wrap_text=False, vertical="top")
        if t and not t.startswith(" ") and not t.startswith("  ") and t[0].isalpha() and t[0].isupper():
            c.font = Font(bold=True, size=12)


def write_mode_sheet(wb, model_name, mode):
    ws = wb.create_sheet(title=mode["name"])
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="F2CB00")
    label_fill = PatternFill("solid", fgColor="EEEEEE")

    ws["A1"] = "Model:";  ws["B1"] = model_name
    ws["A2"] = "Mode:";   ws["B2"] = mode["name"]
    ws["A3"] = "Type:";   ws["B3"] = "WTG"
    for c in ("A1", "A2", "A3"):
        ws[c].font = bold
        ws[c].fill = label_fill

    # Wind speed header row 4. Cell A4 stays blank.
    for ci, ws_mps in enumerate(WIND_SPEEDS, start=2):
        cell = ws.cell(row=4, column=ci, value=ws_mps)
        cell.font = bold
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="right")

    # Frequency + Lw grid from row 5 down.
    for ri, freq in enumerate(FREQS_HZ, start=5):
        ws.cell(row=ri, column=1, value=freq).font = bold
        for ci, ws_mps in enumerate(WIND_SPEEDS, start=2):
            lw = lw_for(mode["lw_at_8ms"], ws_mps, ri - 5)
            if lw is None:
                continue
            c = ws.cell(row=ri, column=ci, value=lw)
            c.alignment = Alignment(horizontal="right")
            c.number_format = "0.0"

    # Column widths so it reads nicely.
    ws.column_dimensions["A"].width = 14
    for ci in range(2, 2 + len(WIND_SPEEDS)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 9


def main():
    wb = openpyxl.Workbook()
    write_readme(wb.active)
    model_name = "EXAMPLE-WTG-4.5MW"
    for mode in MODES:
        write_mode_sheet(wb, model_name, mode)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
