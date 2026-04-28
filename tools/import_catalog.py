#!/usr/bin/env python
"""Import bundled WTG / BESS / Auxiliary catalog data from .xlsx files in
`Import data/` and emit a TypeScript seed module the web app loads on first
launch.

Two file formats are supported:

  - "WTG-style" (one workbook per model, one sheet per mode):
      A1 = "Model:",  B1 = model name
      A2 = "Mode:",   B2 = mode name
      A3 = "Type:",   B3 = "WTG" / "BESS" / "Auxiliary"   (optional row)
      Wind-speed header row (B onward)
      Frequency × wind-speed grid below

  - "Flat" (one row per (model, mode); single sheet):
      Headers: Model | Mode | Type | <freq1> | <freq2> | ...
      Each row gives the spectrum for that (model, mode) combo.

Both formats land in the same TypeScript shape (CatalogEntry per model with
one or more CatalogModeData per mode).

Usage:
    python tools/import_catalog.py
"""

import json
import os
import re
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
IMPORT_DIR = ROOT / 'Import data'
OUTPUT_FILE = ROOT / 'web' / 'src' / 'lib' / 'seedCatalog.ts'

KIND_MAP = {
    'WTG': 'wtg',
    'WIND TURBINE': 'wtg',
    'WINDTURBINE': 'wtg',
    'BESS': 'bess',
    'BATTERY': 'bess',
    'INVERTER': 'auxiliary',
    'TRANSFORMER': 'auxiliary',
    'AUXILIARY': 'auxiliary',
    'AUX': 'auxiliary',
    'OTHER': 'auxiliary',
}


def slugify(name: str) -> str:
    s = re.sub(r'[^a-z0-9]+', '-', (name or '').lower()).strip('-')
    return s or 'unknown'


def is_third_octave(frequencies):
    """Median consecutive ratio < 1.5 ⇒ third-octave (≈1.26); else octave (≈2)."""
    fs = [f for f in frequencies if f and f > 0]
    if len(fs) < 2:
        return False
    ratios = [fs[i] / fs[i - 1] for i in range(1, len(fs)) if fs[i - 1] > 0]
    if not ratios:
        return False
    ratios.sort()
    return ratios[len(ratios) // 2] < 1.5


def extract_rotor_diameter(name: str):
    """Pull rotor diameter from model strings like 'V163 4.5 MW' or 'GE 3.6-137'."""
    if not name:
        return None
    # Vestas: V<diameter> ...
    m = re.search(r'\bV[- ]?(\d{2,3})\b', name)
    if m:
        return int(m.group(1))
    # GE / Siemens style: ...-<diameter>
    m = re.search(r'-(\d{2,3})\b', name)
    if m:
        return int(m.group(1))
    return None


def parse_wtg_workbook(wb, fname):
    """One workbook → one CatalogEntry with one ModeData per sheet."""
    entry = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        a1 = ws.cell(1, 1).value
        a2 = ws.cell(2, 1).value
        if a1 not in ('Model:', 'Model') or a2 not in ('Mode:', 'Mode'):
            print(f'  skip sheet {sheet_name!r}: unexpected layout', file=sys.stderr)
            continue
        model = (ws.cell(1, 2).value or '').strip()
        mode = (ws.cell(2, 2).value or '').strip()
        if not model or not mode:
            print(f'  skip sheet {sheet_name!r}: missing model/mode', file=sys.stderr)
            continue

        # Detect optional Type row.
        a3 = ws.cell(3, 1).value
        if a3 in ('Type:', 'Type'):
            kind_str = (ws.cell(3, 2).value or 'WTG').strip()
            wind_row, data_start = 4, 5
        else:
            kind_str = 'WTG'
            wind_row, data_start = 3, 4
        kind = KIND_MAP.get(kind_str.upper(), 'wtg')

        # Wind-speed header.
        wind_speeds = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(wind_row, c).value
            if v is None:
                break
            try:
                wind_speeds.append(float(v))
            except (TypeError, ValueError):
                break

        # Frequency × wind-speed grid.
        frequencies = []
        spectra = {str(int(w) if w == int(w) else w): [] for w in wind_speeds}
        for r in range(data_start, ws.max_row + 1):
            f = ws.cell(r, 1).value
            if f is None:
                continue
            try:
                freq = float(f)
            except (TypeError, ValueError):
                continue
            frequencies.append(freq)
            for i, w in enumerate(wind_speeds):
                v = ws.cell(r, 2 + i).value
                key = str(int(w) if w == int(w) else w)
                spectra[key].append(float(v) if v is not None else 0.0)

        band_system = 'oneThirdOctave' if is_third_octave(frequencies) else 'octave'

        mode_data = {
            'name': mode,
            'bandSystem': band_system,
            'frequencies': frequencies,
            'spectra': spectra,
            'windSpeeds': wind_speeds,
        }

        if entry is None:
            entry = {
                'id': slugify(model),
                'kind': kind,
                'displayName': model,
                'defaultMode': mode,
                'modes': [],
                'source': fname,
                'origin': 'seed',
            }
            rotor = extract_rotor_diameter(model)
            if rotor is not None:
                entry['rotorDiameterM'] = rotor
        entry['modes'].append(mode_data)

    return [entry] if entry else []


def parse_flat_workbook(wb, fname):
    """Flat layout: one row per (model, mode); columns 4+ are frequencies."""
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if len(headers) < 4:
        return []

    # Frequency columns start at index 3 (column D).
    freq_cols = []
    for i, h in enumerate(headers):
        if i < 3:
            continue
        try:
            f = float(h)
            freq_cols.append((i + 1, f))
        except (TypeError, ValueError):
            continue
    if not freq_cols:
        return []
    frequencies = [f for (_, f) in freq_cols]
    band_system = 'oneThirdOctave' if is_third_octave(frequencies) else 'octave'

    entries = {}
    for r in range(2, ws.max_row + 1):
        model = ws.cell(r, 1).value
        mode = ws.cell(r, 2).value
        type_str = ws.cell(r, 3).value
        if not model or mode is None:
            continue
        model = str(model).strip()
        mode = str(mode).strip()
        kind_str = (type_str or 'AUXILIARY').strip()
        kind = KIND_MAP.get(kind_str.upper(), 'auxiliary')

        spectrum = []
        for col, _ in freq_cols:
            v = ws.cell(r, col).value
            spectrum.append(float(v) if v is not None else 0.0)

        mode_data = {
            'name': mode,
            'bandSystem': band_system,
            'frequencies': frequencies,
            'spectra': {'broadband': spectrum},
        }

        if model not in entries:
            entries[model] = {
                'id': slugify(model),
                'kind': kind,
                'displayName': model,
                'defaultMode': mode,
                'modes': [],
                'source': fname,
                'origin': 'seed',
            }
            if kind == 'auxiliary':
                entries[model]['auxiliaryType'] = kind_str.lower()
        entries[model]['modes'].append(mode_data)

    return list(entries.values())


def parse_workbook(path):
    fname = os.path.basename(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    first = wb[wb.sheetnames[0]]
    a1 = first.cell(1, 1).value
    b1 = first.cell(1, 2).value
    if a1 == 'Model:':
        return parse_wtg_workbook(wb, fname)
    if a1 == 'Model' and isinstance(b1, str) and b1.strip().lower() == 'mode':
        return parse_flat_workbook(wb, fname)
    raise ValueError(f'{fname}: unrecognised layout (A1={a1!r}, B1={b1!r})')


def emit_ts(entries):
    json_blob = json.dumps(entries, indent=2, ensure_ascii=False)
    return f"""// AUTO-GENERATED — do not edit. Regenerate via:
//   python tools/import_catalog.py
//
// Source files in `Import data/` (relative to repo root). To add a new
// model, drop a .xlsx into that folder and re-run the importer.

import type {{ CatalogEntry }} from './types';

export const SEED_CATALOG: CatalogEntry[] = {json_blob};
"""


def main():
    if not IMPORT_DIR.exists():
        print(f'no Import data folder at {IMPORT_DIR}', file=sys.stderr)
        sys.exit(1)
    all_entries = []
    for path in sorted(IMPORT_DIR.iterdir()):
        if path.name.startswith('~$') or path.suffix.lower() != '.xlsx':
            continue
        print(f'== {path.name}')
        try:
            entries = parse_workbook(str(path))
        except Exception as e:
            print(f'  ERROR: {e}', file=sys.stderr)
            continue
        for e in entries:
            n_modes = len(e['modes'])
            n_freqs = len(e['modes'][0]['frequencies']) if e['modes'] else 0
            print(f'  + {e["displayName"]} ({e["kind"]}) — {n_modes} mode(s), {n_freqs} bands')
        all_entries.extend(entries)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(emit_ts(all_entries), encoding='utf-8')
    print(f'\n{len(all_entries)} catalog entries written to {OUTPUT_FILE}')


if __name__ == '__main__':
    main()
